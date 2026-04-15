from __future__ import annotations
import os, sqlite3, datetime as dt, zipfile, shutil
from pathlib import Path
from functools import wraps
from typing import Optional, Dict, Any, List

from flask import Flask, render_template, request, redirect, url_for, flash, session, send_from_directory, abort, send_file
from werkzeug.utils import secure_filename
import openpyxl

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get('PVL_DATA_DIR', str(BASE_DIR / 'data')))
MONTHS_DIR = DATA_DIR / 'months'
BACKUP_DIR = DATA_DIR / 'backups'
DB_PATH = DATA_DIR / 'app.db'

MONTHS = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
    7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
}

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'pvl-web-v1-3-secret-key-change-me')
app.config['MAX_CONTENT_LENGTH'] = 25 * 1024 * 1024
app.config['UPLOAD_EXTENSIONS'] = ['.xlsx', '.xlsm']


def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def ensure_dirs() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    MONTHS_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)


def init_db() -> None:
    ensure_dirs()
    conn = get_db()
    conn.executescript(
        '''
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password TEXT NOT NULL,
            password_changed_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS month_files (
            month INTEGER NOT NULL,
            file_type TEXT NOT NULL,
            path TEXT NOT NULL,
            uploaded_at TEXT NOT NULL,
            PRIMARY KEY (month, file_type)
        );
        CREATE TABLE IF NOT EXISTS stock_snapshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT NOT NULL,
            leche REAL NOT NULL,
            harina REAL NOT NULL,
            notes TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS print_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT NOT NULL,
            month INTEGER NOT NULL,
            cant INTEGER,
            numero INTEGER,
            comite TEXT NOT NULL,
            leche REAL NOT NULL,
            harina REAL NOT NULL,
            recogio INTEGER NOT NULL DEFAULT 0,
            stock_leche_antes REAL,
            stock_harina_antes REAL,
            stock_leche_despues REAL,
            stock_harina_despues REAL,
            source TEXT DEFAULT 'web'
        );
        '''
    )
    existing = conn.execute('SELECT username FROM users WHERE username = ?', ('pvl',)).fetchone()
    if not existing:
        conn.execute(
            'INSERT INTO users(username, password, password_changed_at) VALUES (?, ?, ?)',
            ('pvl', '2026', dt.date.today().isoformat())
        )
    defaults = {
        'admin_password': 'J0el1905',
        'app_version': 'V1.3',
        'author_footer': 'creado por CPC Joel Enrique Clariana Saavedra',
        'last_backup_at': '',
    }
    for k, v in defaults.items():
        if not conn.execute('SELECT 1 FROM settings WHERE key=?', (k,)).fetchone():
            conn.execute('INSERT INTO settings(key, value) VALUES(?, ?)', (k, v))
    conn.commit()
    conn.close()
    preload_month_file_records()
    if not latest_backup_path():
        create_backup('initial')


def preload_month_file_records() -> None:
    conn = get_db()
    now = dt.datetime.now().isoformat(timespec='seconds')
    for month in range(1, 13):
        for file_type in ('pecosa', 'acta'):
            for ext in ('.xlsm', '.xlsx'):
                path = MONTHS_DIR / f'{month:02d}_{file_type}{ext}'
                if path.exists():
                    conn.execute(
                        'INSERT OR REPLACE INTO month_files(month, file_type, path, uploaded_at) VALUES (?, ?, ?, ?)',
                        (month, file_type, str(path), now)
                    )
                    break
    conn.commit()
    conn.close()


def create_backup(reason: str = 'manual') -> Path:
    ensure_dirs()
    stamp = dt.datetime.now().strftime('%Y%m%d_%H%M%S')
    out = BACKUP_DIR / f'backup_{stamp}_{reason}.zip'
    with zipfile.ZipFile(out, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        if DB_PATH.exists():
            zf.write(DB_PATH, arcname='app.db')
        for file in sorted(MONTHS_DIR.glob('*.*')):
            zf.write(file, arcname=f'months/{file.name}')
    backups = sorted(BACKUP_DIR.glob('backup_*.zip'), reverse=True)
    for extra in backups[30:]:
        extra.unlink(missing_ok=True)
    set_setting('last_backup_at', dt.datetime.now().isoformat(timespec='seconds'))
    return out


def latest_backup_path() -> Optional[Path]:
    backups = sorted(BACKUP_DIR.glob('backup_*.zip'), reverse=True)
    return backups[0] if backups else None


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get('user'):
            return redirect(url_for('login'))
        return view(*args, **kwargs)
    return wrapped


def password_expired(username: str) -> bool:
    conn = get_db()
    row = conn.execute('SELECT password_changed_at FROM users WHERE username=?', (username,)).fetchone()
    conn.close()
    if not row:
        return False
    changed = dt.date.fromisoformat(row['password_changed_at'])
    return (dt.date.today() - changed).days >= 30


def get_setting(key: str, default: str = '') -> str:
    conn = get_db()
    row = conn.execute('SELECT value FROM settings WHERE key=?', (key,)).fetchone()
    conn.close()
    return row['value'] if row else default


def set_setting(key: str, value: str) -> None:
    conn = get_db()
    conn.execute('INSERT OR REPLACE INTO settings(key, value) VALUES(?, ?)', (key, value))
    conn.commit()
    conn.close()


def month_enabled(month: int) -> bool:
    return month <= dt.date.today().month


def allowed_file(filename: str) -> bool:
    return Path(filename).suffix.lower() in app.config['UPLOAD_EXTENSIONS']


def get_month_file(month: int, file_type: str) -> Optional[Path]:
    for ext in ('.xlsm', '.xlsx'):
        p = MONTHS_DIR / f'{month:02d}_{file_type}{ext}'
        if p.exists():
            return p
    return None


def normalize_text(value: Any) -> str:
    return ' '.join(str(value or '').upper().split())


def _float(v: Any) -> float:
    try:
        return float(v or 0)
    except Exception:
        return 0.0


def load_month_committees(month: int) -> List[Dict[str, Any]]:
    path = get_month_file(month, 'pecosa') or get_month_file(month, 'acta')
    if not path:
        return []

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = None
    for candidate in ['COMITES', 'base de datos']:
        if candidate in wb.sheetnames:
            sheet = wb[candidate]
            break
    if sheet is None:
        return []

    rows: List[Dict[str, Any]] = []
    headers = [normalize_text(c.value) for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    if 'CANT' in headers:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None for v in row[:3]):
                continue
            rows.append({
                'month': month,
                'cant': row[0],
                'numero': row[1],
                'comite': str(row[2] or '').strip(),
                'dni_presidenta': row[3] if len(row) > 3 else None,
                'presidenta': row[4] if len(row) > 4 else None,
                'dni_almacenera': row[5] if len(row) > 5 else None,
                'almacenera': row[6] if len(row) > 6 else None,
                'beneficiarios': row[7] if len(row) > 7 else None,
                'leche': _float(row[8] if len(row) > 8 else 0),
                'harina': _float(row[9] if len(row) > 9 else 0),
            })
    else:
        for r in range(5, sheet.max_row + 1):
            cant = sheet.cell(r, 2).value
            comite = sheet.cell(r, 3).value
            if cant is None or comite in (None, ''):
                continue
            rows.append({
                'month': month,
                'cant': cant,
                'numero': None,
                'comite': str(comite).strip(),
                'dni_presidenta': sheet.cell(r, 4).value,
                'presidenta': sheet.cell(r, 5).value,
                'dni_almacenera': sheet.cell(r, 6).value,
                'almacenera': sheet.cell(r, 7).value,
                'beneficiarios': sheet.cell(r, 8).value,
                'leche': _float(sheet.cell(r, 9).value),
                'harina': _float(sheet.cell(r, 10).value),
            })
    return rows


def search_committee(query: str, months: List[int]) -> List[Dict[str, Any]]:
    q = normalize_text(query)
    results: List[Dict[str, Any]] = []
    seen = set()
    for month in months:
        for item in load_month_committees(month):
            fields = [normalize_text(item.get('cant')), normalize_text(item.get('numero')), normalize_text(item.get('comite'))]
            if any(q in field for field in fields if field):
                key = (month, item.get('cant'), item.get('comite'))
                if key not in seen:
                    results.append(item)
                    seen.add(key)
    return results


def initial_stock() -> Dict[str, float]:
    conn = get_db()
    snap = conn.execute('SELECT leche, harina FROM stock_snapshots ORDER BY id DESC LIMIT 1').fetchone()
    conn.close()
    if not snap:
        return {'leche': 0.0, 'harina': 0.0}
    return {'leche': float(snap['leche']), 'harina': float(snap['harina'])}


def current_stock() -> Dict[str, float]:
    stock = initial_stock()
    conn = get_db()
    rows = conn.execute('SELECT leche, harina, recogio FROM print_log ORDER BY id').fetchall()
    conn.close()
    leche = stock['leche']
    harina = stock['harina']
    for row in rows:
        if int(row['recogio']) == 1:
            leche -= float(row['leche'])
            harina -= float(row['harina'])
    return {'leche': leche, 'harina': harina}


def today_summary() -> Dict[str, Any]:
    conn = get_db()
    today = dt.date.today().isoformat()
    snap = conn.execute('SELECT leche, harina, created_at FROM stock_snapshots ORDER BY id DESC LIMIT 1').fetchone()
    logs = conn.execute('SELECT * FROM print_log WHERE date(created_at)=? AND recogio=1 ORDER BY id', (today,)).fetchall()
    conn.close()

    stock_ini_leche = float(snap['leche']) if snap else 0.0
    stock_ini_harina = float(snap['harina']) if snap else 0.0
    current_leche = stock_ini_leche
    current_harina = stock_ini_harina
    detail = []
    for row in logs:
        before_l, before_h = current_leche, current_harina
        current_leche -= float(row['leche'])
        current_harina -= float(row['harina'])
        detail.append({
            'hora': row['created_at'][11:19],
            'month': row['month'],
            'cant': row['cant'],
            'numero': row['numero'],
            'comite': row['comite'],
            'leche': float(row['leche']),
            'harina': float(row['harina']),
            'stock_leche_antes': before_l,
            'stock_harina_antes': before_h,
            'stock_leche_despues': current_leche,
            'stock_harina_despues': current_harina,
        })
    return {
        'fecha': today,
        'stock_inicial_leche': stock_ini_leche,
        'stock_inicial_harina': stock_ini_harina,
        'movimientos': detail,
        'stock_final_leche': current_leche,
        'stock_final_harina': current_harina,
    }


@app.context_processor
def inject_globals():
    now = dt.datetime.now()
    return {
        'MONTHS': MONTHS,
        'current_year': now.year,
        'app_version': get_setting('app_version', 'V1.3'),
        'author_footer': get_setting('author_footer', 'creado por CPC Joel Enrique Clariana Saavedra'),
        'now_live': now.strftime('%d/%m/%Y %H:%M:%S')
    }


@app.route('/')
def index():
    if session.get('user'):
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        conn = get_db()
        row = conn.execute('SELECT * FROM users WHERE username=? AND password=?', (username, password)).fetchone()
        conn.close()
        if row:
            session['user'] = username
            session['admin_ok'] = False
            if password_expired(username):
                session['force_password_change'] = True
                return redirect(url_for('change_password'))
            return redirect(url_for('dashboard'))
        flash('Usuario o contraseña incorrecta.', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    force = session.get('force_password_change', False)
    if request.method == 'POST':
        current = request.form.get('current_password', '')
        new = request.form.get('new_password', '')
        confirm = request.form.get('confirm_password', '')
        conn = get_db()
        row = conn.execute('SELECT * FROM users WHERE username=?', (session['user'],)).fetchone()
        if not row or row['password'] != current:
            flash('La contraseña actual no coincide.', 'danger')
        elif len(new) < 4:
            flash('La nueva contraseña debe tener al menos 4 caracteres.', 'danger')
        elif new != confirm:
            flash('La confirmación no coincide.', 'danger')
        else:
            conn.execute('UPDATE users SET password=?, password_changed_at=? WHERE username=?', (new, dt.date.today().isoformat(), session['user']))
            conn.commit()
            conn.close()
            create_backup('password_change')
            session['force_password_change'] = False
            flash('Contraseña actualizada correctamente.', 'success')
            return redirect(url_for('dashboard'))
        conn.close()
    return render_template('change_password.html', force=force)


@app.route('/dashboard', methods=['GET', 'POST'])
@login_required
def dashboard():
    selected_months = [m for m in range(1, 13) if month_enabled(m)]
    results = []
    query = ''
    chosen_months = selected_months
    if request.method == 'POST':
        query = request.form.get('query', '').strip()
        chosen_months = sorted([int(m) for m in request.form.getlist('months') if m.isdigit()])
        allowed = [m for m in chosen_months if month_enabled(m)]
        chosen_months = allowed or selected_months
        if query:
            results = search_committee(query, chosen_months)
    return render_template(
        'dashboard.html',
        results=results,
        query=query,
        selected_months=chosen_months,
        stock=current_stock(),
        months_status={m: {'enabled': month_enabled(m), 'has_pecosa': bool(get_month_file(m, 'pecosa')), 'has_acta': bool(get_month_file(m, 'acta'))} for m in range(1, 13)},
        latest_backup=latest_backup_path().name if latest_backup_path() else None,
        last_backup_at=get_setting('last_backup_at', '')
    )


@app.route('/print/<int:month>/<int:cant>')
@login_required
def print_committee(month: int, cant: int):
    item = next((x for x in load_month_committees(month) if int(x.get('cant') or 0) == cant), None)
    if not item:
        abort(404)
    return render_template('print_committee.html', item=item)


@app.route('/confirm-pickup', methods=['POST'])
@login_required
def confirm_pickup():
    month = int(request.form['month'])
    cant = int(request.form['cant'])
    numero = request.form.get('numero') or None
    comite = request.form['comite']
    leche = float(request.form['leche'])
    harina = float(request.form['harina'])
    recogio = 1 if request.form.get('recogio') == 'si' else 0
    stock = current_stock()
    stock_before_l = stock['leche']
    stock_before_h = stock['harina']
    stock_after_l = stock_before_l - leche if recogio else stock_before_l
    stock_after_h = stock_before_h - harina if recogio else stock_before_h

    conn = get_db()
    conn.execute(
        '''INSERT INTO print_log(
               created_at, month, cant, numero, comite, leche, harina, recogio,
               stock_leche_antes, stock_harina_antes, stock_leche_despues, stock_harina_despues, source
           ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (
            dt.datetime.now().isoformat(timespec='seconds'), month, cant, numero, comite, leche, harina, recogio,
            stock_before_l, stock_before_h, stock_after_l, stock_after_h, 'web'
        )
    )
    conn.commit()
    conn.close()
    create_backup('pickup')
    flash('Movimiento registrado correctamente.', 'success')
    return redirect(url_for('dashboard'))


@app.route('/consolidado-dia')
@login_required
def consolidado_dia():
    return render_template('consolidado.html', data=today_summary())


@app.route('/admin/check', methods=['POST'])
@login_required
def admin_check():
    password = request.form.get('admin_password', '')
    if password == get_setting('admin_password', 'J0el1905'):
        session['admin_ok'] = True
        return redirect(url_for('admin_months'))
    flash('Contraseña de administración incorrecta.', 'danger')
    return redirect(url_for('dashboard'))


@app.route('/admin/months', methods=['GET', 'POST'])
@login_required
def admin_months():
    if not session.get('admin_ok'):
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'upload':
            month = int(request.form['month'])
            file_type = request.form['file_type']
            up = request.files.get('file')
            if month_enabled(month) and up and allowed_file(up.filename):
                ext = Path(secure_filename(up.filename)).suffix.lower()
                out = MONTHS_DIR / f'{month:02d}_{file_type}{ext}'
                for p in MONTHS_DIR.glob(f'{month:02d}_{file_type}.*'):
                    p.unlink(missing_ok=True)
                up.save(out)
                conn = get_db()
                conn.execute(
                    'INSERT OR REPLACE INTO month_files(month, file_type, path, uploaded_at) VALUES (?, ?, ?, ?)',
                    (month, file_type, str(out), dt.datetime.now().isoformat(timespec='seconds'))
                )
                conn.commit()
                conn.close()
                create_backup(f'upload_{month}_{file_type}')
                flash(f'Archivo de {file_type} para {MONTHS[month]} cargado correctamente.', 'success')
            else:
                flash('No se pudo cargar el archivo. Revisa el mes habilitado y la extensión.', 'danger')
        elif action == 'stock':
            leche = float(request.form.get('stock_leche') or 0)
            harina = float(request.form.get('stock_harina') or 0)
            notes = request.form.get('notes', '')
            conn = get_db()
            conn.execute('INSERT INTO stock_snapshots(created_at, leche, harina, notes) VALUES(?,?,?,?)',
                         (dt.datetime.now().isoformat(timespec='seconds'), leche, harina, notes))
            conn.commit()
            conn.close()
            create_backup('stock')
            flash('Stock actualizado correctamente.', 'success')
    status = {m: {'enabled': month_enabled(m), 'pecosa': get_month_file(m, 'pecosa'), 'acta': get_month_file(m, 'acta')} for m in range(1,13)}
    conn = get_db()
    stock_history = conn.execute('SELECT * FROM stock_snapshots ORDER BY id DESC LIMIT 10').fetchall()
    conn.close()
    backups = sorted(BACKUP_DIR.glob('backup_*.zip'), reverse=True)[:10]
    return render_template('admin_months.html', status=status, stock=current_stock(), stock_history=stock_history, backups=backups)


@app.route('/backup/latest')
@login_required
def backup_latest():
    latest = latest_backup_path()
    if not latest:
        abort(404)
    return send_file(latest, as_attachment=True, download_name=latest.name)


@app.route('/backup/create', methods=['POST'])
@login_required
def backup_create():
    create_backup('manual')
    flash('Respaldo creado correctamente.', 'success')
    return redirect(url_for('admin_months'))


@app.route('/files/<path:filename>')
@login_required
def files(filename: str):
    return send_from_directory(MONTHS_DIR, filename, as_attachment=True)


init_db()

if __name__ == '__main__':
    host = os.environ.get('HOST', '0.0.0.0')
    port = int(os.environ.get('PORT', '5050'))
    app.run(host=host, port=port, debug=False)
